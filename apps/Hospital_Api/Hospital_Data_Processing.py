# -*- coding: utf-8 -*-
from Connect_Mongodb import *
import xlrd
import re
import time
import math
import difflib
import jieba
from openpyxl import load_workbook

import sys

reload(sys)
sys.setdefaultencoding('utf-8')


"""
################################################获取所有医院信息###############################################
"""


# 获取医生的信息
def Hospital_Info():
    client = get_db('YunshiDB_Hospital_Info')
    posts = client['YunShiDB']
    data = posts.find({}, {'_id': 0}).limit(10).skip(8)
    dict_dir = {}
    dict_dir['code'] = 1
    dict_dir['content'] = []
    for td in data:
        dict_dir['content'].append(td)
    return dict_dir


"""
################################################医院信息精确匹配###############################################
"""


# 根据Excel文件提供的医院名和省份的列表,精确查找医院信息
def get_accurate_match_info(param, posts, cur):
    filename = "t1u" + str(param["id"])
    path = r"/mnt/newdata/Yunshi_Args/%s.xlsx" % filename
    data = xlrd.open_workbook(path)
    table = data.sheets()[0]
    nrows = table.nrows
    # 当excel文件中只有一行或者小于一行数据时，则表示没有需要处理的数据，执行结束了
    if nrows <= 1:
        param["status"] = 1
        update_status(param, posts, cur)
    # list_dir = []
    start_time = time.time()
    for row in range(1, nrows):
        args_dict = {}
        args_dict['Province'] = table.cell(row, 6).value
        args_dict['City'] = table.cell(row, 7).value
        args_dict['County'] = table.cell(row, 8).value
        args_dict['HCOName'] = table.cell(row, 1).value
        args_dict['ByName1'] = table.cell(row, 2).value
        args_dict['ByName2'] = table.cell(row, 3).value
        args_dict['ByName3'] = table.cell(row, 4).value
        args_dict['ByName4'] = table.cell(row, 5).value
        try:
            result = accurate_match_hospital_info(args_dict)
        except:
            continue
        # list_dir.append(result)
        try:
            Input_Accurate_Match_Excel(result, path, row+1)
        except:
            continue
        end_time = time.time()
        if end_time - start_time >= 5.0:
            # 统计当前数据处理进度
            param["status"] = int(row * 1.0 / (nrows - 1) * 100) if int(row * 1.0 / (nrows - 1) * 100) > 3 else 3
            param["finished"] = row - 1
            # 更新数据库中status值
            start_time = update_status(param, posts, cur)
    # Input_Accurate_Match_Excel(list_dir, path)


# 根据省份和医院名信息,精确查找医院信息
def accurate_match_hospital_info(args_dict):
    client = get_db('YunshiDB_Hospital_Info')
    posts = client['YunShiDB']
    # 如果某一行只有序列号,其他值全空的话,返回空字典
    if not args_dict.values():
        return {}
    # 存储数据
    dict_dir = {}
    if args_dict["Province"] and args_dict["City"] and args_dict["County"]:
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+省+市+区
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},
                                        {"Dynamic_Information.Address_Information.City": args_dict['City']},
                                        {"Dynamic_Information.Address_Information.County": args_dict['County']},
                                        {"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    elif(args_dict["Province"] and args_dict["City"] and not args_dict["County"]):
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+省+市
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},
                                        {"Dynamic_Information.Address_Information.City": args_dict['City']},
                                        {"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    elif(args_dict["Province"] and args_dict["County"] and not args_dict["City"]):
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+省+区
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},
                                        {"Dynamic_Information.Address_Information.County": args_dict['County']},
                                        {"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    elif(args_dict["Province"] and not args_dict["County"] and not args_dict["City"]):
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+省
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},
                                        {"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    elif(not args_dict["Province"] and args_dict["County"] and not args_dict["City"]):
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+区
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.County": args_dict['County']},
                                        {"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    elif (not args_dict["Province"] and not args_dict["County"] and args_dict["City"]):
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+市
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.City": args_dict['City']},
                                        {"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    else:
        # 医院名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名
        data = posts.find_one({"$and": [{"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
    # 若按医院名+省+市规则匹配成功，则匹配等级1
    if data:
        dict_dir['Mapping_Level'] = 1
    # 若医院名(excel)与数据库中医院名未匹配成功，则医院名(excel)与数据库中医院别名进行精确匹配,筛选条件为医院名+省+市
    if not data:
        for j in range(1, 5):
            Alias = "Basic_Information.MedicalOrgAlias%d" % j
            data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']}, {"Dynamic_Information.Address_Information.City": args_dict['City']}, {Alias: args_dict['HCOName']}]}, {"_id": 0})
            if data:
                dict_dir['Mapping_Level'] = 1
                break
    # 若医院名(excel)与数据库中医院别名未匹配成功，则医院别名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+省+市
    if not data:
        for j in range(1, 5):
            ByName = 'ByName%d' % j
            if args_dict[ByName]:
                data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']}, {"Dynamic_Information.Address_Information.City": args_dict['City']},{"Basic_Information.MedicalOrgName": args_dict[ByName]}]}, {"_id": 0})
                # 若按医院别名 + 省 + 市规则匹配成功，则匹配等级2
                if data:
                    dict_dir['Mapping_Level'] = 2
                    break
    # 若医院别名(excel)与数据库中医院名未匹配成功，则医院别名(excel)与数据库中医院别名进行精确匹配,筛选条件为医院名+省+市
    if not data:
        for j in range(1, 5):
            ByName = "ByName%d" % j
            Alias = "Basic_Information.MedicalOrgAlias%d" % j
            if args_dict[ByName]:
                data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']}, {"Dynamic_Information.Address_Information.City": args_dict['City']}, {Alias: args_dict[ByName]}]}, {"_id": 0})
                if data:
                    dict_dir['Mapping_Level'] = 2
                    break
    # 若按医院名+省+市规则未匹配成功，则医院名(excel)与数据库中医院名进行精确匹配,筛选条件改为医院名+省
    if not data:
        data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},{"Basic_Information.MedicalOrgName": args_dict['HCOName']}]}, {"_id": 0})
        # 若医院名(excel)与数据库中医院名未匹配成功，则医院名(excel)与数据库中医院别名进行精确匹配,筛选条件为医院名+市
        if not data:
            for j in range(1, 5):
                Alias = "Basic_Information.MedicalOrgAlias%d" % j
                data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},{Alias: args_dict['HCOName']}]}, {"_id": 0})
                if data:
                    break
        # 若按医院名 + 省规则匹配成功，则匹配等级3
        if data:
            dict_dir['Mapping_Level'] = 3
    # 若医院名(excel)与数据库中医院别名未匹配成功，则医院别名(excel)与数据库中医院名进行精确匹配,筛选条件为医院名+省
    if not data:
        for j in range(1, 5):
            ByName = "ByName%d" % j
            if args_dict[ByName]:
                data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},{"Basic_Information.MedicalOrgName": args_dict[ByName]}]}, {"_id": 0})
                # 若按医院别名 + 省规则匹配成功，则匹配等级4
                if data:
                    dict_dir['Mapping_Level'] = 4
                    break
    # 若医院别名(excel)与数据库中医院名未匹配成功，则医院别名(excel)与数据库中医院别名进行精确匹配,筛选条件为医院名+省+市
    if not data:
        for j in range(1, 5):
            ByName = "ByName%d" % j
            Alias = "Basic_Information.MedicalOrgAlias%d" % j
            if args_dict[ByName]:
                data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": args_dict['Province']},{Alias: args_dict[ByName]}]}, {"_id": 0})
                # 若按医院别名 + 省规则匹配成功，则匹配等级4
                if data:
                    dict_dir['Mapping_Level'] = 4
                    break
    # 若匹配成功，则将匹配成功的医院信息返回并将数据保存到excel文件
    if data:
        # 将一个json文件的信息整合在一个dict_dir中,方便将数据存储到excel文件中，首先获取医院的基本信息
        dict_dir.update(data['Basic_Information'])
        # 获取医院发展情况
        Business_Information = data['Dynamic_Information'][0]['Business_Information']
        dict_dir.update(Business_Information)
        # 获取医院地址信息
        Address_Information = data['Dynamic_Information'][0]['Address_Information']
        dict_dir.update(Address_Information)
        # 获取医院床位信息
        Bed_Information = data['Dynamic_Information'][0]['Bed_Information']
        dict_dir.update(Bed_Information)
        # 获取医院管理者信息
        Administrative_Information = data['Dynamic_Information'][0]['Administrative_Information']
        dict_dir.update(Administrative_Information)
        # 获取医院医生信息
        Doctor_Information = data['Dynamic_Information'][0]['Doctor_Information']
        dict_dir.update(Doctor_Information)
        # 获取医院患者信息
        Patient_Information = data['Dynamic_Information'][0]['Patient_Information']
        dict_dir.update(Patient_Information)
        # 获取医院的ID
        dict_dir['MedicalOrgID'] = data['MedicalOrgID']
        dict_dir['Features'] = data['Features']
        # 获取医院信息状态码
        dict_dir.update(data['Flags'])
        # 获取医院信息修改情况
        dict_dir.update(data['Record'])
        # Input_Accurate_Match_Excel(dict_dir, row)
    return dict_dir


# 将数据导入到Accurate_Match_Info表中
def Input_Accurate_Match_Excel(dict_dir, path, row):
    ws = load_workbook(path)
    wb = ws.active
    wb.title = 'sheet1'
    if dict_dir:
        for col in range(9, 114):
            key = re.findall(r'(\([a-zA-Z0-9/_]+\))', wb.cell(row=1, column=col).value)
            if key and dict_dir. has_key(key[0].strip('(').strip(')')):
                wb.cell(row=row, column=col, value=dict_dir[key[0].strip('(').strip(')')])
    ws.save(path)
    print "We Have Inserted %d Hospital Infomation Sucessfully" % row


"""
################################################医院信息模糊匹配###############################################
"""


# 根据Excel文件提供的医院名和省份的列表,模糊查找医院信息
def get_blur_match_info(param, posts, cur):
    # list_dir = []
    filename = "t2u" + str(param["id"])
    path = r"/mnt/newdata/Yunshi_Args/%s.xlsx" % filename
    data = xlrd.open_workbook(path)
    table = data.sheets()[0]
    nrows = table.nrows
    # 当excel文件中只有一行或者小于一行数据时，则表示没有需要处理的数据，执行结束了
    if nrows <= 1:
        param["status"] = 1
        update_status(param, posts, cur)
    start_time = time.time()
    for row in range(1, nrows):
        dict_args = {}
        dict_args['Province'] = table.cell(row, 6).value
        dict_args['City'] = table.cell(row, 7).value
        dict_args['County'] = table.cell(row, 8).value
        dict_args['HCO_Name'] = table.cell(row, 1).value
        try:
            result_info = compute_similarity(dict_args)
        except:
            continue
        # 如果通过相似度算法进行模糊匹配，未找到相似的医院，则跳出本次循环
        # list_dir.append(result_info)
        try:
            Input_Blur_Match_Excel(result_info, path, row+1)
        except:
            continue
        end_time = time.time()
        if end_time - start_time >= 5.0:
            # 统计当前数据处理进度
            param["status"] = int(row * 1.0 / (nrows - 1) * 100) if int(row * 1.0 / (nrows - 1) * 100) > 3 else 3
            # 统计当前已经处理的条数
            param["finished"] = row - 1
            # 更新数据库中status值
            start_time = update_status(param, posts, cur)
    # 将结果保存到excel中
    # Input_Blur_Match_Excel(list_dir, path)


# 计算两个医院名的相似度,一个医院名来自参数传入,另一个医院名从数据库获取
def compute_similarity(args_info):
    # 连接数据库
    client = get_db('YunshiDB_Hospital_Info')
    posts = client['YunShiDB']
    if not args_info.values():
        return []
    if args_info['Province'] and args_info['City'] and args_info['County']:
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为省+市+区
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info['Province']},
                                    {"Dynamic_Information.Address_Information.City": args_info['City']},
                                    {"Dynamic_Information.Address_Information.County": args_info['County']}]},
                          {'_id': 0})
    elif(args_info['Province'] and args_info['City'] and not args_info['County']):
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为省+市
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info['Province']},
                                    {"Dynamic_Information.Address_Information.City": args_info['City']}]},
                          {'_id': 0})
    elif (args_info['Province'] and not args_info['City'] and args_info['County']):
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为省+区
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info['Province']},
                                    {"Dynamic_Information.Address_Information.County": args_info['County']}]},
                          {'_id': 0})
    elif (not args_info['Province'] and args_info['City'] and args_info['County']):
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为市+区
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.City": args_info['City']},
                                    {"Dynamic_Information.Address_Information.County": args_info['County']}]},
                          {'_id': 0})
    elif (args_info['Province'] and not args_info['City'] and not args_info['County']):
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为省
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info['Province']}]},
                          {'_id': 0})
    elif (not args_info['Province'] and args_info['City'] and not args_info['County']):
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为市
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.City": args_info['City']}]},
                          {'_id': 0})
    elif (not args_info['Province'] and not args_info['City'] and args_info['County']):
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配，筛选条件为区
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.County": args_info['County']}]},
                          {'_id': 0})
    else:
        # 首先将医院名进行结巴分词，然后与数据库医院名进行初步的模糊匹配
        data = posts.find({'_id': 0})
    '''
    # 若根据医院名(excel)与数据库医院名模糊匹配未找到数据,则说明数据库中没有该数据
    if not list(data):
        return []
    '''
    # 保存相似度最大值的医院名、省份信息以及相似度值
    list_dir = []
    for td in data:
        # 保存参数
        args_dict = {}
        # 获取省份信息
        args_dict['Province'] = td['Dynamic_Information'][0]['Address_Information']['Province']
        # print args_dict['Province'], args_info['Province']
        # 若两个医院的省份不同,则说明不可能是同一家医院
        if args_dict['Province'] != args_info['Province']:
            continue
        # 获取医院ID
        args_dict['FeaturesId'] = td["Features"]
        # 获取医院名
        args_dict['HCOName'] = td['Basic_Information']['MedicalOrgName']
        args_dict['ByName1'] = td['Basic_Information']['MedicalOrgAlias1']
        args_dict['ByName2'] = td['Basic_Information']['MedicalOrgAlias2']
        args_dict['ByName3'] = td['Basic_Information']['MedicalOrgAlias3']
        args_dict['ByName4'] = td['Basic_Information']['MedicalOrgAlias4']
        # 计算医院名(excel)和数据库中医院名的相似度
        HCOName_Similarity_Value = difflib.SequenceMatcher(None, args_info['HCO_Name'].encode('utf8'), args_dict['HCOName'].encode('utf8')).ratio()
        # 如果数据库有医院别名1，则计算医院名(excel)和数据库中医院别名1的相似度，若数据库没有医院别名1，则相似度值设为0
        if args_dict['ByName1']:
            ByName1_Similarity_Value = difflib.SequenceMatcher(None, args_info['HCO_Name'].encode('utf8'), args_dict['ByName1'].encode('utf8')).ratio()
        else:
            ByName1_Similarity_Value = 0
        # 如果数据库有医院别名2，则计算医院名(excel)和数据库中医院别名2的相似度，若数据库没有医院别名2，则相似度值设为0
        if args_dict['ByName2']:
            ByName2_Similarity_Value = difflib.SequenceMatcher(None, args_info['HCO_Name'].encode('utf8'), args_dict['ByName2'].encode('utf8')).ratio()
        else:
            ByName2_Similarity_Value = 0
        # 如果数据库有医院别名3，则计算医院名(excel)和数据库中医院别名3的相似度，若数据库没有医院别名3，则相似度值设为0
        if args_dict['ByName3']:
            ByName3_Similarity_Value = difflib.SequenceMatcher(None, args_info['HCO_Name'].encode('utf8'), args_dict['ByName3'].encode('utf8')).ratio()
        else:
            ByName3_Similarity_Value = 0
        # 如果数据库有医院别名4，则计算医院名(excel)和数据库中医院别名4的相似度，若数据库没有医院别名4，则相似度值设为0
        if args_dict['ByName4']:
            ByName4_Similarity_Value = difflib.SequenceMatcher(None, args_info['HCO_Name'].encode('utf8'), args_dict['ByName4'].encode('utf8')).ratio()
        else:
            ByName4_Similarity_Value = 0
        # 获得医院名(excel)和数据库医院名相似度局部最大值
        Max_Similarity_Value = max(HCOName_Similarity_Value, ByName1_Similarity_Value, ByName2_Similarity_Value, ByName3_Similarity_Value, ByName4_Similarity_Value)
        # 相似度值最大的医院ID、医院名以及相似度值保存在list_dir中
        list_tmp = []
        # 保存医院ID
        list_tmp.append(args_dict['FeaturesId'])
        # 保存相似度最大值
        list_tmp.append(Max_Similarity_Value)
        # 保存相似度最大的医院名
        if HCOName_Similarity_Value == Max_Similarity_Value:
            if Revise_Similarity(args_info['HCO_Name'], args_dict['HCOName']):
                list_tmp.append(args_dict['HCOName'])
                list_dir.append(list_tmp)
                continue
        if ByName1_Similarity_Value == Max_Similarity_Value:
            if Revise_Similarity(args_info['HCO_Name'], args_dict['ByName1']):
                list_tmp.append(args_dict['ByName1'])
                list_dir.append(list_tmp)
                continue
        if ByName2_Similarity_Value == Max_Similarity_Value:
            if Revise_Similarity(args_info['HCO_Name'], args_dict['ByName2']):
                list_tmp.append(args_dict['ByName2'])
                list_dir.append(list_tmp)
                continue
        if ByName3_Similarity_Value == Max_Similarity_Value:
            if Revise_Similarity(args_info['HCO_Name'], args_dict['ByName3']):
                list_tmp.append(args_dict['ByName3'])
                list_dir.append(list_tmp)
                continue
        if ByName4_Similarity_Value == Max_Similarity_Value:
            if Revise_Similarity(args_info['HCO_Name'], args_dict['ByName4']):
                list_tmp.append(args_dict['ByName4'])
                list_dir.append(list_tmp)
    # 列表按照相似度值由大到小排序
    list_dir.sort(lambda x, y: cmp(x[1], y[1]), reverse=True)
    # 通过省份信息、相似度计算以及医院名校验（Revise_Similarity函数）三步，模糊匹配出来的医院按照相似度值由大到小排序输出前5个医院信息，如果结果数小于或等于5个，则全部输出，如果结果数大于5个取前5个
    if len(list_dir) <= 5:
        return list_dir
    return list_dir[:5]


# 当获取到相似度最大值医院信息后,通过结巴分词再次确认是否是同一家医院
def Revise_Similarity(Unmatch_Name,Match_Name):
    # print Unmatch_Name,Match_Name
    # 状态码，标记两个医院是否是同一家医院 ,默认标记为1
    flag = 1
    if not Match_Name or not Unmatch_Name:
        return 0
    # 匹配两个医院名字符串中的大写数字，如果同时存在数字并且不相等或者只有一个医院名有数字则说明不可能是同一家医院
    Unmatch_Upper_Digit = re.findall(r"一|二|三|四|五|六|七|八|九|十", Unmatch_Name.encode('utf-8'))
    Match_Upper_Digit = re.findall(r"一|二|三|四|五|六|七|八|九|十", Match_Name.encode('utf-8'))
    if Unmatch_Upper_Digit and Match_Upper_Digit and Unmatch_Upper_Digit != Match_Upper_Digit:
        flag *= 0
        return flag
    if not Unmatch_Upper_Digit and Match_Upper_Digit or Unmatch_Upper_Digit and not Match_Upper_Digit:
        flag *= 0
        return flag
    # 匹配两个医院名字符串中的小写数字，如果同时存在数字并且不相等则说明不可能是同一家医院
    Unmatch_Lower_Digit = re.findall(r"([\d]+)", Unmatch_Name.encode('utf-8'))
    Match_Lower_Digit = re.findall(r"([\d]+)", Match_Name.encode('utf-8'))
    if Unmatch_Lower_Digit and Match_Lower_Digit and Unmatch_Lower_Digit==Match_Lower_Digit:
        flag *= 0
        return flag
    # 或者只有一个医院名有数字则说明不可能是同一家医院
    if not Unmatch_Lower_Digit and Match_Lower_Digit or Unmatch_Lower_Digit and not Match_Lower_Digit:
        flag *= 0
        return flag
    # 如果两个医院名都含有省份但是省份不相等则说明不可能是同一家医院
    # 同理，如果两个医院名都含有市但是市不相等则说明不可能是同一家医院，如果两个医院名都含有区但是区不相等则说明不可能是同一家医院
    # 对两个医院名进行结巴分词
    Unmatch_Name_List = list(jieba.cut(Unmatch_Name))
    Match_Name_List = list(jieba.cut(Match_Name))
    # 结巴分词后，得到两个词向量，对两个词向量进行相似度计算，flag标记两个医院是否相似，相似标记为1，不相似标记为0
    for td in Unmatch_Name_List:
        if u'省' in td:
            for tr in Match_Name_List:
                if (u'省' in tr) and (td != tr):
                    flag *= 0
                    continue
        if u'市' in td:
            for tr in Match_Name_List:
                if (u'市' in tr) and (td != tr):
                    flag *= 0
                    continue
        if (u'区' in td) and (u'社区' not in td):
            for tr in Match_Name_List:
                if (u'区' in tr) and (u'社区' not in td) and (td != tr):
                    flag *= 0
                    continue
        if u'县' in td:
            for tr in Match_Name_List:
                if (u'县' in tr) and (td != tr):
                    flag *= 0
                    continue
        if u'镇' in td:
            for tr in Match_Name_List:
                if (u'镇' in tr) and (td != tr):
                    flag *= 0
                    continue
        if u'社区' in td:
            for tr in Match_Name_List:
                if (u'社区' in tr) and (td != tr):
                    flag *= 0
                    continue
    return flag


# 根据省份和医院名信息,精确查找医院信息
def get_blur_match_hospital_info(Province, Hospital_Name, row):
    client = get_db('YunshiDB_Hospital_Info')
    posts = client['YunShiDB']
    data = posts.find_one({"$and": [{"Dynamic_Information.Address_Information.Province": Province}, {"Basic_Information.MedicalOrgName": Hospital_Name}]}, {"_id": 0})
    #print json.dumps(data)
    dict_dir = {}
    if data:
        # 将一个json文件的信息整合在一个dict_dir中,方便将数据存储到excel文件中，首先获取医院的基本信息
        dict_dir = data['Basic_Information']
        # 获取医院发展情况
        Business_Information = data['Dynamic_Information'][0]['Business_Information']
        dict_dir.update(Business_Information)
        # 获取医院地址信息
        Address_Information = data['Dynamic_Information'][0]['Address_Information']
        dict_dir.update(Address_Information)
        # 获取医院床位信息
        Bed_Information = data['Dynamic_Information'][0]['Bed_Information']
        dict_dir.update(Bed_Information)
        # 获取医院管理者信息
        Administrative_Information = data['Dynamic_Information'][0]['Administrative_Information']
        dict_dir.update(Administrative_Information)
        # 获取医院医生信息
        Doctor_Information = data['Dynamic_Information'][0]['Doctor_Information']
        dict_dir.update(Doctor_Information)
        # 获取医院患者信息
        Patient_Information = data['Dynamic_Information'][0]['Patient_Information']
        dict_dir.update(Patient_Information)
        # 获取医院的ID
        dict_dir['MedicalOrgID'] = data['MedicalOrgID']
        dict_dir['Features'] = data['Features']
        # 获取医院信息状态码
        dict_dir.update(data['Flags'])
        # 获取医院信息修改情况
        dict_dir.update(data['Record'])
        Input_Blur_Match_Excel(dict_dir, row)
    print data['MedicalOrgID']
    return dict_dir


# 将数据插入到Blur_Match_Info表文件中
def Input_Blur_Match_Excel(list_dir, path, row):
    ws = load_workbook(path)
    wb = ws.active
    wb.title = 'Sheet1'
    # 将结果保存到excel文件中，用row表示文件行数
    # row = 2
    # 将结果保存到excel文件中
    # 将结果保存到excel文件中,定义count来记录结果数量
    count = 1
    # 遍历模糊匹配前五个结果
    for info in list_dir:
        args_list = ['Features%d' % count, 'Similarity%d' % count, 'MedicalOrgName%d' % count]
        dict_dir = dict(zip(args_list, info))
        for col in range(10, 25):
            key = re.findall(r'(\([a-zA-Z0-9/_]+\))', wb.cell(row=1, column=col).value)
            if key and dict_dir. has_key(key[0].strip('(').strip(')')):
                wb.cell(row=row, column=col, value=dict_dir[key[0].strip('(').strip(')')])
        count += 1
    ws.save(path)
    print "We Have Inserted %d Hospital Infomation Sucessfully" % row


"""
################################################根据医院ID查找医院信息###############################################
"""


# 根据医院ID查找相关医院信息
def get_hospital_info_byId(param, posts_mysql, cur):
    client = get_db('YunshiDB_Hospital_Info')
    posts = client['YunShiDB']
    filename = "t4u" + str(param["id"])
    path = r"/mnt/newdata/Yunshi_Args/%s.xlsx" % filename
    data = xlrd.open_workbook(path)
    table = data.sheets()[0]
    nrows = table.nrows
    # 当excel文件中只有一行或者小于一行数据时，则表示没有需要处理的数据，执行结束了
    if nrows <= 1:
        param["status"] = 1
        update_status(param, posts_mysql, cur)
    # list_dir = []
    start_time = time.time()
    for row in range(1, nrows):
        dict_tmp = {}
        try:
            HcoId = table.cell(row, 1).value
        except:
            continue
        try:
            data = posts.find_one({"Features": HcoId}, {"_id": 0})
        except:
            continue
        if data:
            # 将一个json文件的信息整合在一个dict_dir中,方便将数据存储到excel文件中，首先获取医院的基本信息
            dict_tmp = data['Basic_Information']
            # 获取医院发展情况
            Business_Information = data['Dynamic_Information'][0]['Business_Information']
            dict_tmp.update(Business_Information)
            # 获取医院地址信息
            Address_Information = data['Dynamic_Information'][0]['Address_Information']
            dict_tmp.update(Address_Information)
            # 获取医院床位信息
            Bed_Information = data['Dynamic_Information'][0]['Bed_Information']
            dict_tmp.update(Bed_Information)
            # 获取医院管理者信息
            Administrative_Information = data['Dynamic_Information'][0]['Administrative_Information']
            dict_tmp.update(Administrative_Information)
            # 获取医院医生信息
            Doctor_Information = data['Dynamic_Information'][0]['Doctor_Information']
            dict_tmp.update(Doctor_Information)
            # 获取医院患者信息
            Patient_Information = data['Dynamic_Information'][0]['Patient_Information']
            dict_tmp.update(Patient_Information)
            # 获取医院的ID
            dict_tmp['MedicalOrgID'] = data['MedicalOrgID']
            dict_tmp['Features'] = data['Features']
            # 获取医院信息状态码
            dict_tmp.update(data['Flags'])
            # 获取医院信息修改情况
            dict_tmp.update(data['Record'])
            # Input_ByHospitalId_Excel(dict_tmp, row+1)
        # list_dir.append(dict_tmp)
        try:
            Input_ByHospitalId_Excel(dict_tmp, path, row+1)
        except:
            continue
        end_time = time.time()
        if end_time - start_time >= 5.0:
            # 统计当前数据处理进度
            param["status"] = int(row * 1.0 / (nrows - 1) * 100) if int(row * 1.0 / (nrows - 1) * 100) > 3 else 3
            param["finished"] = row - 1
            # 更新数据库中status值
            start_time = update_status(param, posts_mysql, cur)
        print "The FeatureId %s Has Found" % HcoId
    # Input_ByHospitalId_Excel(list_dir, path)


# 将数据保存到HcoId_Match_Info文件中
def Input_ByHospitalId_Excel(dict_dir, path, row):
    ws = load_workbook(path)
    wb = ws.active
    wb.title = 'Sheet1'
    # row = 2
    if dict_dir:
        for col in range(3, 106):
            key = re.findall(r'(\([a-zA-Z0-9/_]+\))', wb.cell(row=1, column=col).value)
            if key and dict_dir. has_key(key[0].strip('(').strip(')')):
                wb.cell(row=row, column=col, value=dict_dir[key[0].strip('(').strip(')')])
    row += 1
    ws.save(path)
    print "We Have Inserted %d Hospital Infomation Sucessfully" % row


"""
################################################根据输入的省份、市、区查找医院信息###############################################
"""


# 根据提供的省、市、区信息查找相关医院信息
def get_all_thing(args_dict, posts, cur):
    data, total = Get_Hospital_Info(args_dict)
    dict_dir = {"code": 1, "msg": "success", "content": []}
    row = 1
    start_time = time.time()
    for td in data:
        dict_tmp = {}
        # 将一个json文件的信息整合在一个dict_dir中,方便将数据存储到excel文件中，首先获取医院的基本信息
        dict_tmp = td['Basic_Information']
        # 获取医院地址信息
        Address_Information = td['Dynamic_Information'][0]['Address_Information']
        dict_tmp.update(Address_Information)
        # 获取医院的ID
        dict_tmp['MedicalOrgID'] = td['MedicalOrgID']
        dict_tmp['Features'] = td['Features']
        # 获取医院信息修改情况
        dict_dir["content"].append(dict_tmp)
        end_time = time.time()
        if end_time - start_time >= 5.0:
            # 统计当前数据处理进度
            args_dict["status"] = int(row * 1.0 / int(total) * 100) if int(row * 1.0 / int(total) * 100) > 3 else 3
            # 更新数据库中status值
            start_time = update_status_tmp(args_dict, posts, cur)
        row += 1
    dict_dir["pagesize"] = int(args_dict["pagesize"])
    dict_dir["page"] = int(args_dict["page"])
    dict_dir["totalpage"] = int(math.ceil(int(total)*1.0/int(args_dict['pagesize'])))
    return dict_dir


# 连接数据获取相关的医院的数据
def Get_Hospital_Info(args_info):
    client = get_db("YunshiDB_Hospital_Info")
    posts = client["YunShiDB"]
    data_num = (int(args_info["page"])-1)*int(args_info["pagesize"])
    # 对医院名进行模糊匹配
    # 两个参数都不为空
    if args_info["Param1"] and args_info["Param2"]:
        print "####################1########################"
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param1"], "$options": "i"}},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param2"], "$options": "i"}}]},
                          {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
        total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param1"], "$options": "i"}},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param2"], "$options": "i"}}]}).count()
    # Param1存在和Param2为空
    elif (args_info["Param1"] and not args_info["Param2"]):
        print "####################2########################"
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param1"], "$options": "i"}}]},
                          {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
        print type(data)
        total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param1"], "$options": "i"}}]}).count()
    # Param1为空和Param2存在
    elif (args_info["Param2"] and not args_info["Param1"]):
        print "####################3########################"
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param2"], "$options": "i"}}]},
                          {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
        print type(data)
        total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                    {"Basic_Information.MedicalOrgName": {"$regex": args_info["Param2"], "$options": "i"}}]}).count()
    # Param1为空和Param2为空
    else:
        print "####################4########################"
        data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]}]},
                          {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
        total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                    {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                    {"Dynamic_Information.Address_Information.County": args_info["County"]}]}).count()
        # 若按照医院名去匹配没查询找数据,则用医院别名去匹配
    if not data:
        # 总共4个别名,循环一下就行
        for i in range(1, 5):
            if args_info["Param1"] and args_info["Param2"]:
                hospital_name = "Basic_Information.MedicalOrgAlias%d" % i
                data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                            {hospital_name: {"$regex": args_info["Param1"], "$options": "i"}},
                                            {hospital_name: {"$regex": args_info["Param2"], "$options": "i"}}]},
                                  {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
                total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                            {hospital_name: {"$regex": args_info["Param1"], "$options": "i"}},
                                            {hospital_name: {"$regex": args_info["Param2"], "$options": "i"}}]}).count()
            elif (args_info["Param1"] and not args_info["Param2"]):
                data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                            {hospital_name: {"$regex": args_info["Param1"], "$options": "i"}}]},
                                  {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
                total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                            {hospital_name: {"$regex": args_info["Param1"], "$options": "i"}}]}).count()
            elif (args_info["Param2"] and not args_info["Param1"]):
                data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                            {hospital_name: {"$regex": args_info["Param2"], "$options": "i"}}]},
                                  {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
                total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]},
                                            {hospital_name: {"$regex": args_info["Param2"], "$options": "i"}}]}).count()
            else:
                data = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]}]},
                                  {"_id": 0}).limit(int(args_info["pagesize"])).skip(data_num)
                total_num = posts.find({"$and": [{"Dynamic_Information.Address_Information.Province": args_info["Province"]},
                                            {"Dynamic_Information.Address_Information.City": args_info["City"]},
                                            {"Dynamic_Information.Address_Information.County": args_info["County"]}]}).count()
            # 如果匹配到数据，直接跳出循环
            if data:
                break
    return data, total_num


# 通过省、市、区查询到的医院信息保存在excel中
def Input_Hospital_Info_Excel(dict_dir):
    path = r'/mnt/newdata/Yunshi_Args/Hospital_Info.xlsx'
    ws = load_workbook(path)
    wb = ws.active
    wb.title = 'Sheet1'
    nrows = wb.rows
    row = len(nrows) + 1
    for col in range(1, 104):
        key = re.findall(r'(\([a-zA-Z0-9/_]+\))', wb.cell(row=1, column=col).value)
        if key and dict_dir. has_key(key[0].strip('(').strip(')')):
            wb.cell(row=row, column=col, value=dict_dir[key[0].strip('(').strip(')')])
        row += 1
    ws.save(path)
    print "We Have Inserted %d Hospital Infomation Sucessfully" % row


"""
################################################插入最新的医院数据##########################################
"""


# 将Excel文件的数据导入到MongoDB
def Insert_Hospital_Info(param, posts, cur):
    filename = "t5u" + str(param["id"])
    path = r"/mnt/newdata/Yunshi_Args/%s.xlsx" % filename
    ws = load_workbook(path)
    wb = ws.active
    wb.title = 'Sheet1'
    nrows = wb.rows
    # 当excel文件中只有一行或者小于一行数据时，则表示没有需要处理的数据，执行结束了
    if len(list(nrows)) <= 1:
        param["status"] = 1
        update_status(param, posts, cur)
    # list_dir = []
    start_time = time.time()
    for row in range(2, len(list(nrows)) + 1):
        dict_dir = {}
        if wb.cell(row=row, column=1).value:
            dict_dir['MedicalOrgID'] = wb.cell(row=row, column=1).value.strip('{').strip('}')
        else:
            dict_dir['MedicalOrgID'] = None
        dict_dir['Features'] = wb.cell(row=row, column=2).value
        # 医院的基本信息
        dict_dir['Basic_Information'] = {}
        dict_dir['Basic_Information']['MedicalOrgName'] = wb.cell(row=row, column=9).value
        dict_dir['Basic_Information']['MedicalOrgAlias1'] = wb.cell(row=row, column=14).value
        dict_dir['Basic_Information']['MedicalOrgAlias2'] = wb.cell(row=row, column=15).value
        dict_dir['Basic_Information']['MedicalOrgAlias3'] = wb.cell(row=row, column=16).value
        dict_dir['Basic_Information']['MedicalOrgAlias4'] = wb.cell(row=row, column=17).value
        dict_dir['Basic_Information']['OrganizationID'] = wb.cell(row=row, column=103).value
        dict_dir['Basic_Information']['HospitalGrade'] = wb.cell(row=row, column=10).value
        dict_dir['Basic_Information']['HospitalLevel'] = wb.cell(row=row, column=11).value
        dict_dir['Basic_Information']['MedicalOrgSubType'] = wb.cell(row=row, column=20).value
        dict_dir['Basic_Information']['Telephone'] = wb.cell(row=row, column=13).value
        dict_dir['Basic_Information']['MedicalOrgURL'] = wb.cell(row=row, column=19).value
        dict_dir['Basic_Information']['State'] = wb.cell(row=row, column=33).value
        dict_dir['Basic_Information']['FoundingTime'] = wb.cell(row=row, column=31).value
        dict_dir['Basic_Information']['GSPLicenseID'] = wb.cell(row=row, column=35).value
        dict_dir['Basic_Information']['ApprovedDepartment'] = wb.cell(row=row, column=29).value
        dict_dir['Basic_Information']['Comprehensive/specialized_hospital'] = wb.cell(row=row, column=22).value
        dict_dir['Basic_Information']['ApprovedDate'] = wb.cell(row=row, column=30).value
        dict_dir['Basic_Information']['ExpiryDate'] = wb.cell(row=row, column=32).value
        dict_dir['Basic_Information']['MedicalOrgDesp'] = wb.cell(row=row, column=18).value
        dict_dir['Basic_Information']['AdministrationCode'] = wb.cell(row=row, column=3).value
        dict_dir['Basic_Information']['BusinessType'] = wb.cell(row=row, column=23).value
        dict_dir['Basic_Information']['Note'] = wb.cell(row=row, column=21).value
        dict_dir['Basic_Information']['Sort'] = wb.cell(row=row, column=42).value
        dict_dir['Basic_Information']['AnestheticDrugBusinessQualification'] = wb.cell(row=row, column=47).value
        dict_dir['Basic_Information']['ProfitNature'] = wb.cell(row=row, column=86).value
        dict_dir['Basic_Information']['Reason'] = wb.cell(row=row, column=34).value
        # 医院的动态信息
        dict_dir['Dynamic_Information'] = []
        dict_tmp = {}
        # 医院地址信息
        dict_tmp['Address_Information'] = {}
        dict_tmp['Address_Information']['Province'] = wb.cell(row=row, column=5).value
        dict_tmp['Address_Information']['City'] = wb.cell(row=row, column=6).value
        dict_tmp['Address_Information']['County'] = wb.cell(row=row, column=7).value
        dict_tmp['Address_Information']['Town'] = wb.cell(row=row, column=8).value
        dict_tmp['Address_Information']['Address'] = wb.cell(row=row, column=12).value
        dict_tmp['Address_Information']['WarehouseAddress'] = wb.cell(row=row, column=24).value
        dict_tmp['Address_Information']['PostCode'] = wb.cell(row=row, column=4).value
        # 医院行政信息
        dict_tmp['Administrative_Information'] = {}
        dict_tmp['Administrative_Information']['LegalRepresentative'] = wb.cell(row=row, column=88).value
        dict_tmp['Administrative_Information']['OrgCEO'] = wb.cell(row=row, column=72).value
        dict_tmp['Administrative_Information']['OrgCQO'] = wb.cell(row=row, column=73).value
        # 医院医生信息
        dict_tmp['Doctor_Information'] = {}
        dict_tmp['Doctor_Information']['Workforce'] = wb.cell(row=row, column=92).value
        dict_tmp['Doctor_Information']['LicensedDoctor'] = wb.cell(row=row, column=90).value
        dict_tmp['Doctor_Information']['LicensedAssistantDoctor'] = wb.cell(row=row, column=91).value
        dict_tmp['Doctor_Information']['DoctorsPulmMedical'] = wb.cell(row=row, column=54).value
        dict_tmp['Doctor_Information']['DoctorsEndocrineMedical'] = wb.cell(row=row, column=56).value
        dict_tmp['Doctor_Information']['DoctorsNerveMedical'] = wb.cell(row=row, column=58).value
        dict_tmp['Doctor_Information']['DoctorsMedical'] = wb.cell(row=row, column=59).value
        dict_tmp['Doctor_Information']['DoctorsGeneralSurgical'] = wb.cell(row=row, column=75).value
        dict_tmp['Doctor_Information']['DoctorsBurnSurgical'] = wb.cell(row=row, column=76).value
        dict_tmp['Doctor_Information']['DoctorsSurgical'] = wb.cell(row=row, column=77).value
        # 医院患者信息
        dict_tmp['Patient_Information'] = {}
        dict_tmp['Patient_Information']['YearlyPatients'] = wb.cell(row=row, column=67).value
        dict_tmp['Patient_Information']['YearlyOutpatients'] = wb.cell(row=row, column=50).value
        dict_tmp['Patient_Information']['YearlyPatientsMedical'] = wb.cell(row=row, column=64).value
        dict_tmp['Patient_Information']['YearlyPatientsSurgical'] = wb.cell(row=row, column=66).value
        dict_tmp['Patient_Information']['YearlyPatientsPediatrics'] = wb.cell(row=row, column=61).value
        dict_tmp['Patient_Information']['YearlyPatientsGynaObs'] = wb.cell(row=row, column=62).value
        dict_tmp['Patient_Information']['YearlyPatientsTradChinese'] = wb.cell(row=row, column=63).value
        dict_tmp['Patient_Information']['YearlyPatientsOthers'] = wb.cell(row=row, column=65).value
        dict_tmp['Patient_Information']['LeaveHospital'] = wb.cell(row=row, column=60).value
        dict_tmp['Patient_Information']['PatientsInHospital'] = wb.cell(row=row, column=71).value
        # 医院床位信息
        dict_tmp['Bed_Information'] = {}
        dict_tmp['Bed_Information']['Beds'] = wb.cell(row=row, column=37).value
        dict_tmp['Bed_Information']['BedsEndocrineMedical'] = wb.cell(row=row, column=55).value
        dict_tmp['Bed_Information']['BedsNerveMedical'] = wb.cell(row=row, column=57).value
        dict_tmp['Bed_Information']['BedsPulmMedical'] = wb.cell(row=row, column=53).value
        dict_tmp['Bed_Information']['BedsMedical'] = wb.cell(row=row, column=52).value
        dict_tmp['Bed_Information']['BedsPediatrics'] = wb.cell(row=row, column=39).value
        dict_tmp['Bed_Information']['BedsGynaObs'] = wb.cell(row=row, column=41).value
        dict_tmp['Bed_Information']['BedsDerm'] = wb.cell(row=row, column=68).value
        dict_tmp['Bed_Information']['BedsENT'] = wb.cell(row=row, column=40).value
        dict_tmp['Bed_Information']['BedsStom'] = wb.cell(row=row, column=46).value
        dict_tmp['Bed_Information']['BedsGeneralPractice'] = wb.cell(row=row, column=70).value
        dict_tmp['Bed_Information']['BedsConsumption'] = wb.cell(row=row, column=43).value
        dict_tmp['Bed_Information']['BedsPsy'] = wb.cell(row=row, column=44).value
        dict_tmp['Bed_Information']['BedsCont'] = wb.cell(row=row, column=36).value
        dict_tmp['Bed_Information']['BedsSurgical'] = wb.cell(row=row, column=74).value
        dict_tmp['Bed_Information']['BedsLocal'] = wb.cell(row=row, column=38).value
        dict_tmp['Bed_Information']['BedsEye'] = wb.cell(row=row, column=79).value
        dict_tmp['Bed_Information']['BedsConv'] = wb.cell(row=row, column=45).value
        dict_tmp['Bed_Information']['BedsBeauty'] = wb.cell(row=row, column=85).value
        dict_tmp['Bed_Information']['BedsPreventionHealth'] = wb.cell(row=row, column=87).value
        dict_tmp['Bed_Information']['BedsSports'] = wb.cell(row=row, column=89).value
        dict_tmp['Bed_Information']['BedsOccu'] = wb.cell(row=row, column=93).value
        dict_tmp['Bed_Information']['BedsChineseWestern'] = wb.cell(row=row, column=94).value
        dict_tmp['Bed_Information']['BedsChinese'] = wb.cell(row=row, column=96).value
        dict_tmp['Bed_Information']['BedsOnco'] = wb.cell(row=row, column=97).value
        dict_tmp['Bed_Information']['BedsOthers'] = wb.cell(row=row, column=69).value
        # 医院发展情况信息
        dict_tmp['Business_Information'] = {}
        dict_tmp['Business_Information']['RegisteredCapital'] = wb.cell(row=row, column=101).value
        dict_tmp['Business_Information']['ExpenseDrugCharges'] = wb.cell(row=row, column=82).value
        dict_tmp['Business_Information']['RevenueOutpatients'] = wb.cell(row=row, column=48).value
        dict_tmp['Business_Information']['RevenueInHospital'] = wb.cell(row=row, column=98).value
        dict_tmp['Business_Information']['RevenueWesternOutpatients'] = wb.cell(row=row, column=49).value
        dict_tmp['Business_Information']['RevenueChineseOutpatients'] = wb.cell(row=row, column=51).value
        dict_tmp['Business_Information']['WesternFee'] = wb.cell(row=row, column=78).value
        dict_tmp['Business_Information']['DrugFee'] = wb.cell(row=row, column=80).value
        dict_tmp['Business_Information']['RevenueDrug'] = wb.cell(row=row, column=81).value
        dict_tmp['Business_Information']['ChineseFee'] = wb.cell(row=row, column=95).value
        dict_tmp['Business_Information']['RevenueWesternInHospital'] = wb.cell(row=row, column=99).value
        dict_tmp['Business_Information']['RevenueChineseInHospital'] = wb.cell(row=row, column=100).value
        dict_tmp['Business_Information']['RevenueAll'] = wb.cell(row=row, column=102).value
        dict_tmp['Business_Information']['BusinessType'] = wb.cell(row=row, column=23).value
        dict_tmp['Business_Information']['BusinessScope'] = wb.cell(row=row, column=83).value
        dict_dir['Dynamic_Information'].append(dict_tmp)
        # 数据创建日期
        dict_dir['Record'] = {}
        dict_dir['Record']['CreateDate'] = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        dict_dir['Record']['DataBaseNum'] = wb.cell(row=row, column=27).value
        dict_dir['Record']['DataFrom'] = wb.cell(row=row, column=28).value
        # 数据修改信息
        dict_dir['Flags'] = {}
        dict_dir['Flags']['DataFlag'] = wb.cell(row=row, column=26).value
        try:
            Import_Hospital_Info(dict_dir)
        except:
            continue
        # list_dir.append(dict_dir)
        end_time = time.time()
        if end_time - start_time >= 5.0:
            # 统计当前数据处理进度
            param["status"] = int((row-1) * 1.0 / (len(list(nrows)) - 1) * 100) if int((row-1) * 1.0 / (len(list(nrows)) - 1) * 100) > 3 else 3
            param["finished"] = row - 1
            # 更新数据库中status值
            start_time = update_status(param, posts, cur)
    # Import_Hospital_Info(list_dir)


# 将数据导入到数据库中
def Import_Hospital_Info(dict_dir):
    client = get_db('YunshiDB_Hospital_Info')
    posts = client['YunshiDB']
    posts.insert(dict_dir)
    print "Insert Hospital Information Successfully"


"""
################################################插入最新的医院数据End##########################################
"""


# 更新数据库中status值
def update_status(param, posts, cur):
    sql = "UPDATE dev.execlog SET status= %(status)s,progress=%(finished)s WHERE id=%(id)s"
    posts.execute(sql, param)
    print "We Have Updated The Status %d" % param["status"]
    cur.commit()
    start_time = time.time()
    return start_time


# 更新数据库中status值
def update_status_tmp(param, posts, cur):
    sql = "UPDATE dev.execlog SET status= %(status)s WHERE id=%(id)s"
    posts.execute(sql, param)
    print "We Have Updated The Status %d" % param["status"]
    cur.commit()
    start_time = time.time()
    return start_time
