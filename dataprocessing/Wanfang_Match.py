# -*- coding: utf-8 -*-
"""
Created on Tue Dec 20 20:30:12 2016

@author: liuwy

从万方数据库中查找医生的论文信息

"""
from openpyxl import load_workbook
import sys
from pymongo import MongoClient
import json

reload(sys)
sys.setdefaultencoding('utf8')

'''
##################################万方数据库查找相关医生#############################
'''
'''
conn = MongoClient("192.168.7.24",27017)
global db
db = conn.WanFang_Doctor_Info
'''

# 从excel文件中读取医生信息
def get_info():
    path = r"C:\Users\liuwy\Desktop\data\source.xlsx"
    ws = load_workbook(path)
    wb = ws.active
    wb.title = "Sheet2"
    # 获取excel文件中医生的信息
    nrows = wb.rows
    count = 1
    for row in range(1,len(nrows)+1):
        dict_dir = {}
        dict_dir['code'] =  wb.cell(row = row,column=1).value
        dict_dir['province'] =  wb.cell(row = row,column=2).value
        dict_dir['city'] =  wb.cell(row = row,column=3).value
        dict_dir['institution'] = wb.cell(row = row,column = 4).value
        dict_dir['flag'] =  wb.cell(row = row,column=5).value
        dict_dir['doctor_name'] = wb.cell(row = row,column = 6).value
        dict_dir['professional'] = wb.cell(row = row,column = 7).value
        dict_dir['sex'] = wb.cell(row = row,column = 8).value
        count = match_doctor(dict_dir,count)
        
        print "###################%d######################" % row


# 数据库中匹配医生
def match_doctor(dict_dir,count):
    posts = db['test']
    data = posts.find({},{'doctor_id':1})
    data = posts.find({"$and":[{"name":dict_dir['doctor_name']}]},{"_id":0})
    list_dir = []
    for td in data:
        list_tmp = []
        list_tmp.append(dict_dir['code'])
        list_tmp.append(dict_dir['province'])
        list_tmp.append(dict_dir['city'])
        list_tmp.append(dict_dir['institution'])
        list_tmp.append(dict_dir['flag'])
        list_tmp.append(dict_dir['doctor_name'])
        list_tmp.append(dict_dir['professional'])
        list_tmp.append(dict_dir['sex'])
        list_tmp.append(td['doctor_id'])
        list_tmp.append(td['hospital'])
        list_tmp.append(td['hospital_url'])
        list_dir.append(list_tmp)        
        count += 1
    save_data(list_dir)
        
    print "We Have Matched %s In Database" % dict_dir['doctor_name']
    return count
        
        
# 保存数据
def save_data(list_tmp):
    path = r"C:\Users\liuwy\Desktop\data\result_only_name.xlsx"
    ws = load_workbook(path)
    wb = ws.active
    wb.title = "Sheet1"
    row = len(wb.rows) + 1
    for td in list_tmp:
        # print type(td)
        for i in range(1,12):
            wb.cell(row = row,column=i,value=td[i-1])
        row += 1
    ws.save(path)    
    
    print "Insert %d Data Successfully!!! " % row
    
'''
##################################将万方医学网上医生的论文信息导入到阿里云服务器中#############################
'''
# 方法一，导出json文件后，再导入进去
# 连接本地数据库
localhost_conn = MongoClient("192.168.7.24",27017)
global localhost_db
localhost_db = localhost_conn.WanFang_Doctor_Info


# 连接阿里云数据库
aliyun_conn = MongoClient(host="115.28.77.178",port=27017)
global aliyun_db
aliyun_db = aliyun_conn.wanfang_doctor
#aliyun_db.authenticate("root","Yunshi2016")


def get_data():
    jsonfile = open('wanfang_doctor_info.json','r')
    for td in jsonfile.readlines():
        #print td
        input_database(json.loads(td))
    
def input_database(data):
    global count
    posts = db['WanFang_Doctor_Papers']
    try:
        posts.insert(data)
        print "Import The %d Hsopital Data Sucessfully!!!" % count
        count += 1
    except Exception,e:
        print e
# 方法二，两个数据库直接同步数据
        
global count 
count = 1

# 连接本地数据获取数据
def get_wanfang_data():
    localhost_posts = localhost_db["test"]
    data = localhost_posts.find({},{"_id":0})
    for td in data:
        print "OK"
        Input_Aliyun_Database(td)
        
# 导入到阿里云数据库   
def Input_Aliyun_Database(doctor_info):
    global count
    aliyun_posts = aliyun_db["WanFang_Doctor_Papers"]
    try:
        aliyun_posts.insert(doctor_info)
        print "Import The %d Hsopital Data Sucessfully!!!" % count
        count += 1
    except Exception,e:
        print e
        
        
if __name__ == '__main__':
    #get_info()
    get_wanfang_data()