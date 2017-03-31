# -*- coding: utf-8 -*-
"""
Created on Tue Nov 29 10:48:50 2016

@author: 何超
"""
import re
from openpyxl import load_workbook
from pymongo import MongoClient
import time
import csv
import pymssql

server = r'localhost'
user = r'sa'
password = r'A136668132'
sql_name = r'XMasterData'
global conn
sql_conn = pymssql.connect(server,user, password,sql_name,charset="utf8") 


global db
global count
count = 1

Mongodb_conn = MongoClient("182.92.96.114", 27017)
db = Mongodb_conn.YunshiDB_Hospital_Info
db.authenticate("root", "123456")


# 读写csv文件，并将数据导入到数据库中
def get_csv_data():
    csvfile = file('D:\yunshidata\Hospital_Data_Processing\Medical_org.csv', 'rb')
    readfile = csv.reader(csvfile)
    for line in readfile:
        print line
    csvfile.close() 
    
    
# 读写excel文件，并将数据导入到数据库中
def get_data():
    #data = xlrd.open_workbook(r'D:\yunshidata\Hospital_Data_Processing\Country_Hospital_Info.xlsx')
    path = r'D:\yunshidata\Hospital_Data_Processing\Hospital_Info.xlsx'
    ws = load_workbook(path)
    wb = ws.active
    wb.title = 'Sheet1'
    nrows = wb.rows
    for row in range(2, len(nrows)+1):
        dict_dir = {}
        if wb.cell(row=row, column=1).value:
            dict_dir['MedicalOrgID'] = wb.cell(row=row, column=1).value.strip('{').strip('}')
        else:
            dict_dir['MedicalOrgID'] = ''
        # 医院的特征ID
        dict_dir['Features'] = wb.cell(row=row,column=2).value
        # 医院的基本信息
        dict_dir['Basic_Information'] = {}
        dict_dir['Basic_Information']['MedicalOrgName'] = wb.cell(row=row,column=9).value
        dict_dir['Basic_Information']['MedicalOrgAlias1'] = wb.cell(row=row,column=14).value
        dict_dir['Basic_Information']['MedicalOrgAlias2'] = wb.cell(row=row,column=15).value
        dict_dir['Basic_Information']['MedicalOrgAlias3'] = wb.cell(row=row,column=16).value
        dict_dir['Basic_Information']['MedicalOrgAlias4'] = wb.cell(row=row,column=17).value
        dict_dir['Basic_Information']['OrganizationID'] = wb.cell(row=row,column=103).value
        dict_dir['Basic_Information']['HospitalGrade'] = wb.cell(row=row,column=10).value
        dict_dir['Basic_Information']['HospitalLevel'] = wb.cell(row=row,column=11).value
        dict_dir['Basic_Information']['MedicalOrgSubType'] = wb.cell(row=row,column=20).value
        dict_dir['Basic_Information']['Telephone'] = wb.cell(row=row,column=13).value
        dict_dir['Basic_Information']['MedicalOrgURL'] = wb.cell(row=row,column=19).value
        dict_dir['Basic_Information']['State'] = wb.cell(row=row,column=33).value
        dict_dir['Basic_Information']['FoundingTime'] = wb.cell(row=row,column=31).value
        dict_dir['Basic_Information']['GSPLicenseID'] = wb.cell(row=row,column=35).value
        dict_dir['Basic_Information']['ApprovedDepartment'] = wb.cell(row=row,column=29).value
        dict_dir['Basic_Information']['Comprehensive/specialized_hospital'] = wb.cell(row=row,column=22).value
        dict_dir['Basic_Information']['ApprovedDate'] = wb.cell(row=row,column=30).value
        dict_dir['Basic_Information']['ExpiryDate'] = wb.cell(row=row,column=32).value
        dict_dir['Basic_Information']['MedicalOrgDesp'] = wb.cell(row=row,column=18).value
        dict_dir['Basic_Information']['AdministrationCode'] = wb.cell(row=row,column=3).value
        dict_dir['Basic_Information']['BusinessType'] = wb.cell(row=row,column=23).value
        dict_dir['Basic_Information']['Note'] = wb.cell(row=row,column=21).value
        dict_dir['Basic_Information']['Sort'] = wb.cell(row=row,column=42).value
        dict_dir['Basic_Information']['AnestheticDrugBusinessQualification'] = wb.cell(row=row,column=47).value
        dict_dir['Basic_Information']['ProfitNature'] = wb.cell(row=row,column=86).value
        dict_dir['Basic_Information']['Reason'] = wb.cell(row=row,column=34).value
        # 医院的动态信息
        dict_dir['Dynamic_Information'] = []
        dict_tmp = {}
        # 医院地址信息
        dict_tmp['Address_Information'] ={}
        dict_tmp['Address_Information']['Province'] = wb.cell(row=row,column=5).value
        dict_tmp['Address_Information']['City'] = wb.cell(row=row,column=6).value
        dict_tmp['Address_Information']['County'] = wb.cell(row=row,column=7).value
        dict_tmp['Address_Information']['Town'] = wb.cell(row=row,column=8).value
        dict_tmp['Address_Information']['Address'] = wb.cell(row=row,column=12).value
        dict_tmp['Address_Information']['WarehouseAddress'] = wb.cell(row=row,column=24).value
        dict_tmp['Address_Information']['PostCode'] = wb.cell(row=row,column=4).value
        #医院行政信息
        dict_tmp['Administrative_Information'] = {}
        dict_tmp['Administrative_Information']['LegalRepresentative'] = wb.cell(row=row,column=88).value
        dict_tmp['Administrative_Information']['OrgCEO'] = wb.cell(row=row,column=72).value
        dict_tmp['Administrative_Information']['OrgCQO'] = wb.cell(row=row,column=73).value
        #医院医生信息
        dict_tmp['Doctor_Information'] = {}
        dict_tmp['Doctor_Information']['Workforce'] = wb.cell(row=row,column=92).value
        dict_tmp['Doctor_Information']['LicensedDoctor'] = wb.cell(row=row,column=90).value
        dict_tmp['Doctor_Information']['LicensedAssistantDoctor'] = wb.cell(row=row,column=91).value
        dict_tmp['Doctor_Information']['DoctorsPulmMedical'] = wb.cell(row=row,column=54).value
        dict_tmp['Doctor_Information']['DoctorsEndocrineMedical'] = wb.cell(row=row,column=56).value
        dict_tmp['Doctor_Information']['DoctorsNerveMedical'] = wb.cell(row=row,column=58).value
        dict_tmp['Doctor_Information']['DoctorsMedical'] = wb.cell(row=row,column=59).value
        dict_tmp['Doctor_Information']['DoctorsGeneralSurgical'] = wb.cell(row=row,column=75).value
        dict_tmp['Doctor_Information']['DoctorsBurnSurgical'] = wb.cell(row=row,column=76).value
        dict_tmp['Doctor_Information']['DoctorsSurgical'] = wb.cell(row=row,column=77).value
        #医院患者信息
        dict_tmp['Patient_Information'] = {}
        dict_tmp['Patient_Information']['YearlyPatients'] = wb.cell(row=row,column=67).value
        dict_tmp['Patient_Information']['YearlyOutpatients'] = wb.cell(row=row,column=50).value
        dict_tmp['Patient_Information']['YearlyPatientsMedical'] = wb.cell(row=row,column=64).value
        dict_tmp['Patient_Information']['YearlyPatientsSurgical'] = wb.cell(row=row,column=66).value
        dict_tmp['Patient_Information']['YearlyPatientsPediatrics'] = wb.cell(row=row,column=61).value
        dict_tmp['Patient_Information']['YearlyPatientsGynaObs'] = wb.cell(row=row,column=62).value
        dict_tmp['Patient_Information']['YearlyPatientsTradChinese'] = wb.cell(row=row,column=63).value
        dict_tmp['Patient_Information']['YearlyPatientsOthers'] = wb.cell(row=row,column=65).value
        dict_tmp['Patient_Information']['LeaveHospital'] = wb.cell(row=row,column=60).value
        dict_tmp['Patient_Information']['PatientsInHospital'] = wb.cell(row=row,column=71).value
        #医院床位信息
        dict_tmp['Bed_Information'] = {}
        dict_tmp['Bed_Information']['Beds'] = wb.cell(row=row,column=37).value
        dict_tmp['Bed_Information']['BedsEndocrineMedical'] = wb.cell(row=row,column=55).value
        dict_tmp['Bed_Information']['BedsNerveMedical'] = wb.cell(row=row,column=57).value
        dict_tmp['Bed_Information']['BedsPulmMedical'] = wb.cell(row=row,column=53).value
        dict_tmp['Bed_Information']['BedsMedical'] = wb.cell(row=row,column=52).value
        dict_tmp['Bed_Information']['BedsPediatrics'] = wb.cell(row=row,column=39).value
        dict_tmp['Bed_Information']['BedsGynaObs'] = wb.cell(row=row,column=41).value
        dict_tmp['Bed_Information']['BedsDerm'] = wb.cell(row=row,column=68).value
        dict_tmp['Bed_Information']['BedsENT'] = wb.cell(row=row,column=40).value
        dict_tmp['Bed_Information']['BedsStom'] = wb.cell(row=row,column=46).value
        dict_tmp['Bed_Information']['BedsGeneralPractice'] = wb.cell(row=row,column=70).value
        dict_tmp['Bed_Information']['BedsConsumption'] = wb.cell(row=row,column=43).value
        dict_tmp['Bed_Information']['BedsPsy'] = wb.cell(row=row,column=44).value
        dict_tmp['Bed_Information']['BedsCont'] = wb.cell(row=row,column=36).value
        dict_tmp['Bed_Information']['BedsSurgical'] = wb.cell(row=row,column=74).value
        dict_tmp['Bed_Information']['BedsLocal'] = wb.cell(row=row,column=38).value
        dict_tmp['Bed_Information']['BedsEye'] = wb.cell(row=row,column=79).value
        dict_tmp['Bed_Information']['BedsConv'] = wb.cell(row=row,column=45).value
        dict_tmp['Bed_Information']['BedsBeauty'] = wb.cell(row=row,column=85).value
        dict_tmp['Bed_Information']['BedsPreventionHealth'] = wb.cell(row=row,column=87).value
        dict_tmp['Bed_Information']['BedsSports'] = wb.cell(row=row,column=89).value
        dict_tmp['Bed_Information']['BedsOccu'] = wb.cell(row=row,column=93).value
        dict_tmp['Bed_Information']['BedsChineseWestern'] = wb.cell(row=row,column=94).value
        dict_tmp['Bed_Information']['BedsChinese'] = wb.cell(row=row,column=96).value
        dict_tmp['Bed_Information']['BedsOnco'] = wb.cell(row=row,column=97).value
        dict_tmp['Bed_Information']['BedsOthers'] = wb.cell(row=row,column=69).value
        #医院发展情况信息
        dict_tmp['Business_Information'] = {}
        dict_tmp['Business_Information']['RegisteredCapital'] = wb.cell(row=row,column=101).value
        dict_tmp['Business_Information']['ExpenseDrugCharges'] = wb.cell(row=row,column=82).value
        dict_tmp['Business_Information']['RevenueOutpatients'] = wb.cell(row=row,column=48).value
        dict_tmp['Business_Information']['RevenueInHospital'] = wb.cell(row=row,column=98).value
        dict_tmp['Business_Information']['RevenueWesternOutpatients'] = wb.cell(row=row,column=49).value
        dict_tmp['Business_Information']['RevenueChineseOutpatients'] = wb.cell(row=row,column=51).value
        dict_tmp['Business_Information']['WesternFee'] = wb.cell(row=row,column=78).value
        dict_tmp['Business_Information']['DrugFee'] = wb.cell(row=row,column=80).value
        dict_tmp['Business_Information']['RevenueDrug'] = wb.cell(row=row,column=81).value
        dict_tmp['Business_Information']['ChineseFee'] = wb.cell(row=row,column=95).value
        dict_tmp['Business_Information']['RevenueWesternInHospital'] = wb.cell(row=row,column=99).value
        dict_tmp['Business_Information']['RevenueChineseInHospital'] = wb.cell(row=row,column=100).value
        dict_tmp['Business_Information']['RevenueAll'] = wb.cell(row=row,column=102).value
        dict_tmp['Business_Information']['BusinessType'] = wb.cell(row=row,column=23).value
        dict_tmp['Business_Information']['BusinessScope'] = wb.cell(row=row,column=83).value
        dict_dir['Dynamic_Information'].append(dict_tmp)
        #数据创建日期
        dict_dir['Record'] = {}
        dict_dir['Record']['CreateDate'] = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        dict_dir['Record']['DataBaseNum'] = wb.cell(row=row,column=27).value
        dict_dir['Record']['DataFrom'] = wb.cell(row=row,column=28).value
        #数据修改信息
        dict_dir['Flags'] = {}
        dict_dir['Flags']['DataFlag'] = wb.cell(row=row,column=26).value
        Import_Yunshi_MongoDB(dict_dir)
        print "...................%d......................." % row
        
        
# 将XMasterData数据导入到服务器数据库中
def Get_XMasterData():
    cursor = sql_conn.cursor(as_dict=True) 
    cursor.execute("select * from dbo.Medical_org")
    for td in cursor:
        dict_tmp = {}
        for key,value in td.items():
            field_name = re.findall(r'(\([a-zA-Z0-9/_]+\))', key)[0].strip('(').strip(')')
            dict_tmp[field_name] = value
        # 生成标准的JSON格式数据
        dict_dir = {}
        dict_dir['MedicalOrgID']=dict_tmp['MedicalOrgID']
        #医院的特征ID
        dict_dir['Features'] = dict_tmp['Features']
        #医院的基本信息
        dict_dir['Basic_Information'] = {}
        dict_dir['Basic_Information']['MedicalOrgName'] = dict_tmp['MedicalOrgName']
        dict_dir['Basic_Information']['MedicalOrgAlias1'] = dict_tmp['MedicalOrgAlias1']
        dict_dir['Basic_Information']['MedicalOrgAlias2'] = dict_tmp['MedicalOrgAlias2']
        dict_dir['Basic_Information']['MedicalOrgAlias3'] = dict_tmp['MedicalOrgAlias3']
        dict_dir['Basic_Information']['MedicalOrgAlias4'] = dict_tmp['MedicalOrgAlias4']
        dict_dir['Basic_Information']['OrganizationID'] = dict_tmp['OrganizationID']
        dict_dir['Basic_Information']['HospitalGrade'] = dict_tmp['HospitalGrade']
        dict_dir['Basic_Information']['HospitalLevel'] = dict_tmp['HospitalLevel']
        dict_dir['Basic_Information']['MedicalOrgSubType'] = dict_tmp['MedicalOrgSubType']
        dict_dir['Basic_Information']['Telephone'] = dict_tmp['Telephone']
        dict_dir['Basic_Information']['MedicalOrgURL'] = dict_tmp['MedicalOrgURL']
        dict_dir['Basic_Information']['State'] = dict_tmp['state']
        dict_dir['Basic_Information']['FoundingTime'] = dict_tmp['FoundingTime']
        dict_dir['Basic_Information']['GSPLicenseID'] = dict_tmp['GSPLicenseID']
        dict_dir['Basic_Information']['ApprovedDepartment'] = dict_tmp['ApprovedDepartment']
        dict_dir['Basic_Information']['Comprehensive/specialized_hospital'] = dict_tmp['Comprehensive/specialized_hospital']
        dict_dir['Basic_Information']['ApprovedDate'] = dict_tmp['ApprovedDate']
        dict_dir['Basic_Information']['ExpiryDate'] = dict_tmp['ExpiryDate']
        dict_dir['Basic_Information']['MedicalOrgDesp'] = dict_tmp['MedicalOrgDesp']
        dict_dir['Basic_Information']['AdministrationCode'] = dict_tmp['GBT']
        dict_dir['Basic_Information']['BusinessType'] = dict_tmp['BusinessType']
        dict_dir['Basic_Information']['Note'] = dict_tmp['Note']
        dict_dir['Basic_Information']['Sort'] = dict_tmp['Sort']
        dict_dir['Basic_Information']['AnestheticDrugBusinessQualification'] = dict_tmp['AnestheticDrugBusinessQualification']
        dict_dir['Basic_Information']['ProfitNature'] = dict_tmp['ProfitNature']
        dict_dir['Basic_Information']['Reason'] = dict_tmp['Reason']
        #医院的动态信息
        dict_dir['Dynamic_Information'] = []
        tmp_info = {}
        #医院地址信息
        tmp_info['Address_Information'] ={}
        tmp_info['Address_Information']['Province'] = dict_tmp['Province']
        tmp_info['Address_Information']['City'] = dict_tmp['City']
        tmp_info['Address_Information']['County'] = dict_tmp['County']
        tmp_info['Address_Information']['Town'] = dict_tmp['Town']
        tmp_info['Address_Information']['Address'] = dict_tmp['Address']
        tmp_info['Address_Information']['WarehouseAddress'] = dict_tmp['WarehouseAddress']
        tmp_info['Address_Information']['PostCode'] = dict_tmp['PostCode']
        # 医院管理结构
        tmp_info['Administrative_Information'] = {}
        tmp_info['Administrative_Information']['LegalRepresentative'] = dict_tmp['LegalRepresentative']
        tmp_info['Administrative_Information']['OrgCEO'] = dict_tmp['OrgCEO']
        tmp_info['Administrative_Information']['OrgCQO'] = dict_tmp['OrgCQO']
        #医院医生介绍
        tmp_info['Doctor_Information'] = {}
        tmp_info['Doctor_Information']['Workforce'] = dict_tmp['Workforce']
        tmp_info['Doctor_Information']['LicensedDoctor'] = dict_tmp['LicensedDoctor']
        tmp_info['Doctor_Information']['LicensedAssistantDoctor'] = dict_tmp['LicensedAssistantDoctor']
        tmp_info['Doctor_Information']['DoctorsPulmMedical'] = dict_tmp['DoctorsPulmMedical']
        tmp_info['Doctor_Information']['DoctorsEndocrineMedical'] = dict_tmp['DoctorsEndocrineMedical']
        tmp_info['Doctor_Information']['DoctorsNerveMedical'] = dict_tmp['DoctorsNerveMedical']
        tmp_info['Doctor_Information']['DoctorsMedical'] = dict_tmp['DoctorsMedical']
        tmp_info['Doctor_Information']['DoctorsGeneralSurgical'] = dict_tmp['DoctorsGeneralSurgical']
        tmp_info['Doctor_Information']['DoctorsBurnSurgical'] = dict_tmp['DoctorsBurnSurgical']
        tmp_info['Doctor_Information']['DoctorsSurgical'] = dict_tmp['DoctorsSurgical']
        # 病人的信息
        tmp_info['Patient_Information'] = {}
        tmp_info['Patient_Information']['YearlyPatients'] = dict_tmp['YearlyPatients']
        tmp_info['Patient_Information']['YearlyOutpatients'] = dict_tmp['YearlyOutpatients']
        tmp_info['Patient_Information']['YearlyPatientsMedical'] = dict_tmp['YearlyPatientsMedical']
        tmp_info['Patient_Information']['YearlyPatientsSurgical'] = dict_tmp['YearlyPatientsSurgical']
        tmp_info['Patient_Information']['YearlyPatientsPediatrics'] = dict_tmp['YearlyPatientsPediatrics']
        tmp_info['Patient_Information']['YearlyPatientsGynaObs'] = dict_tmp['YearlyPatientsGynaObs']
        tmp_info['Patient_Information']['YearlyPatientsTradChinese'] = dict_tmp['YearlyPatientsTradChinese']
        tmp_info['Patient_Information']['YearlyPatientsOthers'] = dict_tmp['YearlyPatientsOthers']
        tmp_info['Patient_Information']['LeaveHospital'] = dict_tmp['LeaveHospital']
        tmp_info['Patient_Information']['PatientsInHospital'] = dict_tmp['PatientsInHospital']
        # 医院床位信息        
        tmp_info['Bed_Information'] = {}
        tmp_info['Bed_Information']['Beds'] = dict_tmp['Beds']
        tmp_info['Bed_Information']['BedsEndocrineMedical'] = dict_tmp['BedsEndocrineMedical']
        tmp_info['Bed_Information']['BedsNerveMedical'] = dict_tmp['BedsNerveMedical']
        tmp_info['Bed_Information']['BedsPulmMedical'] = dict_tmp['BedsPulmMedical']
        tmp_info['Bed_Information']['BedsMedical'] = dict_tmp['BedsMedical']
        tmp_info['Bed_Information']['BedsPediatrics'] = dict_tmp['BedsPediatrics']
        tmp_info['Bed_Information']['BedsGynaObs'] = dict_tmp['BedsGynaObs']
        tmp_info['Bed_Information']['BedsDerm'] = dict_tmp['BedsDerm']
        tmp_info['Bed_Information']['BedsENT'] = dict_tmp['BedsENT']
        tmp_info['Bed_Information']['BedsStom'] = dict_tmp['BedsStom']
        tmp_info['Bed_Information']['BedsGeneralPractice'] = dict_tmp['BedsGeneralPractice']
        tmp_info['Bed_Information']['BedsConsumption'] = dict_tmp['BedsConsumption']
        tmp_info['Bed_Information']['BedsPsy'] = dict_tmp['BedsPsy']
        tmp_info['Bed_Information']['BedsCont'] = dict_tmp['BedsCont']
        tmp_info['Bed_Information']['BedsSurgical'] = dict_tmp['BedsSurgical']
        tmp_info['Bed_Information']['BedsLocal'] = dict_tmp['BedsLocal']
        tmp_info['Bed_Information']['BedsEye'] = dict_tmp['BedsEye']
        tmp_info['Bed_Information']['BedsConv'] = dict_tmp['BedsConv']
        tmp_info['Bed_Information']['BedsBeauty'] = dict_tmp['BedsBeauty']
        tmp_info['Bed_Information']['BedsPreventionHealth'] = dict_tmp['BedsPreventionHealth']
        tmp_info['Bed_Information']['BedsSports'] = dict_tmp['BedsSports']
        tmp_info['Bed_Information']['BedsOccu'] = dict_tmp['BedsOccu']
        tmp_info['Bed_Information']['BedsChineseWestern'] = dict_tmp['BedsChineseWestern']
        tmp_info['Bed_Information']['BedsChinese'] = dict_tmp['BedsChinese']
        tmp_info['Bed_Information']['BedsOnco'] = dict_tmp['BedsOnco']
        tmp_info['Bed_Information']['BedsOthers'] = dict_tmp['BedsOthers']
        # 医院发展情况        
        tmp_info['Business_Information'] = {}
        tmp_info['Business_Information']['RegisteredCapital'] = dict_tmp['RegisteredCapital']
        tmp_info['Business_Information']['ExpenseDrugCharges'] = dict_tmp['ExpenseDrugCharges']
        tmp_info['Business_Information']['RevenueOutpatients'] = dict_tmp['RevenueOutpatients']
        tmp_info['Business_Information']['RevenueInHospital'] = dict_tmp['RevenueInHospital']
        tmp_info['Business_Information']['RevenueWesternOutpatients'] = dict_tmp['RevenueWesternOutpatients']
        tmp_info['Business_Information']['RevenueChineseOutpatients'] = dict_tmp['RevenueChineseOutpatients']
        tmp_info['Business_Information']['WesternFee'] = dict_tmp['WesternFee']
        tmp_info['Business_Information']['DrugFee'] = dict_tmp['DrugFee']
        tmp_info['Business_Information']['RevenueDrug'] = dict_tmp['RevenueDrug']
        tmp_info['Business_Information']['ChineseFee'] = dict_tmp['ChineseFee']
        tmp_info['Business_Information']['RevenueWesternInHospital'] = dict_tmp['RevenueWesternInHospital']
        tmp_info['Business_Information']['RevenueChineseInHospital'] = dict_tmp['RevenueChineseInHospital']
        tmp_info['Business_Information']['RevenueAll'] = dict_tmp['RevenueAll']
        tmp_info['Business_Information']['BusinessType'] = dict_tmp['BusinessType']
        tmp_info['Business_Information']['BusinessScope'] = dict_tmp['BusinessScope']
        dict_dir['Dynamic_Information'].append(tmp_info)
        #数据创建日期
        dict_dir['Record'] = {}
        dict_dir['Record']['CreateDate'] = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        dict_dir['Record']['DataBaseNum'] = dict_tmp['DataBaseNum']
        dict_dir['Record']['DataFrom'] = dict_tmp['DataFrom']
        #数据修改信息
        dict_dir['Flags'] = {}
        dict_dir['Flags']['DataFlag'] = dict_tmp['DataFlag']
        Import_Yunshi_MongoDB(dict_dir)
        
        
def Import_Yunshi_MongoDB(data):
    global count 
    posts = db['YunShiDB']
    try:
        posts.insert(data)
        print "Import The %d Hsopital Data Sucessfully!!!" % count
        count += 1
    except Exception,e:
        print e
        print data['MedicalOrgID']



if __name__ == '__main__':
    #get_data()
    #get_csv_data()
    Get_XMasterData()  